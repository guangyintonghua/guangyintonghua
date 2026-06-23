from __future__ import annotations

import argparse
import hashlib
import json
import shutil
import textwrap
from dataclasses import asdict, dataclass
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageColor, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[1]
WORKFLOW_ROOT = ROOT / "桌布自动生图工作流"
PRODUCT_ROOT = WORKFLOW_ROOT / "产品素材"
OUTPUT_ROOT = WORKFLOW_ROOT / "输出"
CACHE_ROOT = OUTPUT_ROOT / "image2缓存"
GENERATED_IMAGE_ROOT = Path.home() / ".codex" / "generated_images"
FONT_PATH = Path(r"C:\Windows\Fonts\msyh.ttc")


PRODUCT_COPY = {
    "1号品": {
        "title": "北欧米白深灰蓝波点流苏桌布",
        "guide_title": "米白蓝灰波点流苏桌布",
        "title_direction": "北欧低饱和波点桌布",
        "house_style": "北欧奶油原木家居风",
        "house_palette": "浅橡木、米白、雾灰、浅灰蓝",
        "visual_traits": "米白底、低饱和深蓝灰波点、流苏边、轻亚麻感",
        "style_tags": ["北欧简约", "低饱和", "波点", "流苏边", "轻亚麻感"],
        "hero_title": "米白蓝灰点",
        "hero_subtitle": "颜色柔和，家用更耐看",
        "pattern_label": "低饱和波点",
        "scene_hint": "适合日常餐桌、茶几、民宿陈设",
    },
    "2号品": {
        "title": "复古深底碎花流苏棉布桌布",
        "guide_title": "深底黄白碎花流苏桌布",
        "title_direction": "复古田园小碎花棉布桌布",
        "house_style": "复古田园木色庭院风",
        "house_palette": "深木色、原木、奶油白、鼠尾草绿、陶土色",
        "visual_traits": "深咖啡底、黄白小碎花、米白流苏花边、薄棉布、自然垂感",
        "style_tags": ["复古田园", "深底碎花", "米白流苏", "棉布印花", "轻薄垂感"],
        "hero_title": "深底黄白碎花",
        "hero_subtitle": "复古田园感更明显，木色餐桌更好搭",
        "pattern_label": "复古碎花更耐看",
        "scene_hint": "适合餐桌、早餐角、边柜、复古家居陈列",
    },
}


@dataclass
class ProductContext:
    product_name: str
    product_dir: Path
    workbook_path: Path
    texture_image: Path
    edge_image: Path
    title: str
    guide_title: str
    title_direction: str
    house_style: str
    house_palette: str
    visual_traits: str
    category: str
    material: str
    sizes: list[str]
    prices: list[str]
    style_tags: list[str]
    hero_title: str
    hero_subtitle: str
    pattern_label: str
    scene_hint: str


@dataclass
class SceneSpec:
    scene_id: str
    bucket: str
    aspect_label: str
    size: tuple[int, int]
    headline: str
    subtitle: str
    prompt: str


def load_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    if FONT_PATH.exists():
        return ImageFont.truetype(str(FONT_PATH), size=size)
    return ImageFont.load_default()


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def reset_dir(path: Path) -> None:
    if path.exists():
        for child in path.iterdir():
            if child.is_file():
                child.unlink()
            else:
                shutil.rmtree(child)
    path.mkdir(parents=True, exist_ok=True)


def wrap_text(text: str, width: int) -> list[str]:
    return textwrap.wrap(text, width=width, break_long_words=False, break_on_hyphens=False)


def wrap_text_pixels(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, max_width: int) -> list[str]:
    if not text:
        return [""]
    lines: list[str] = []
    current = ""
    for char in text:
        trial = current + char
        bbox = draw.textbbox((0, 0), trial, font=font)
        if current and (bbox[2] - bbox[0]) > max_width:
            lines.append(current)
            current = char
        else:
            current = trial
    if current:
        lines.append(current)
    return lines


def text_block_height(line_count: int, font_size: int, line_gap: int) -> int:
    if line_count <= 0:
        return 0
    return line_count * font_size + (line_count - 1) * line_gap


def measure_text_block(
    draw: ImageDraw.ImageDraw,
    *,
    title: str,
    body: str,
    title_font_size: int,
    body_font_size: int,
    title_width_px: int,
    body_width_px: int,
    line_gap: int,
) -> tuple[int, list[str], list[str]]:
    title_font = load_font(title_font_size)
    body_font = load_font(body_font_size)
    title_lines = wrap_text_pixels(draw, title, title_font, title_width_px)
    body_lines = wrap_text_pixels(draw, body, body_font, body_width_px)
    total_height = text_block_height(len(title_lines), title_font_size, line_gap) + 8 + text_block_height(len(body_lines), body_font_size, line_gap)
    return total_height, title_lines, body_lines


def draw_multiline(
    draw: ImageDraw.ImageDraw,
    text: str,
    xy: tuple[int, int],
    font: ImageFont.ImageFont,
    fill: str,
    line_gap: int = 8,
    max_chars: int = 12,
) -> int:
    x, y = xy
    max_width_px = int(font.size * max_chars * 1.05)
    for line in wrap_text_pixels(draw, text, font, max_width_px):
        draw.text((x, y), line, font=font, fill=fill)
        y += font.size + line_gap
    return y


def crop_cover(source: Image.Image, size: tuple[int, int], focus_bottom: bool = False) -> Image.Image:
    target_w, target_h = size
    src_ratio = source.width / source.height
    dst_ratio = target_w / target_h
    if src_ratio > dst_ratio:
        new_h = target_h
        new_w = int(new_h * src_ratio)
    else:
        new_w = target_w
        new_h = int(new_w / src_ratio)
    resized = source.resize((new_w, new_h), Image.Resampling.LANCZOS)
    left = max((new_w - target_w) // 2, 0)
    top = max(new_h - target_h - 40, 0) if focus_bottom else max((new_h - target_h) // 2, 0)
    return resized.crop((left, top, left + target_w, top + target_h))


def clean_edge_source(source: Image.Image) -> Image.Image:
    right = int(source.width * 0.8)
    bottom = int(source.height * 0.9)
    return source.crop((0, 0, right, bottom))


def clean_text(value: object, fallback: str) -> str:
    text = str(value or "").strip()
    if not text:
        return fallback
    if "?" in text and len(text.replace("?", "")) <= 2:
        return fallback
    return text


def build_listing_title(ctx: ProductContext) -> str:
    title = ctx.title.strip()
    if len(title) >= 30:
        return title[:30]

    suffix_parts: list[str] = []
    if ctx.category == "桌布":
        suffix_parts.extend(["餐桌", "茶几", "边柜", "装饰", "盖布", "家用", "布艺"])
    suffix_parts.extend([ctx.material, ctx.category, ctx.guide_title, ctx.title_direction])

    for part in suffix_parts:
        clean_part = str(part).replace("、", "").replace("，", "").replace(" ", "").strip()
        if not clean_part:
            continue
        if clean_part in title:
            continue
        title += clean_part
        if len(title) >= 30:
            return title[:30]
    return title[:30]


def build_guide_title(ctx: ProductContext) -> str:
    title = ctx.guide_title.strip()
    if len(title) >= 15:
        return title[:15]

    suffix_parts: list[str] = ["餐桌", "盖布", "家用", "台布", "茶几"]
    for part in suffix_parts:
        if len(title) >= 15:
            break
        if part in title:
            continue
        title += part
    if len(title) < 15:
        title += "布艺装饰"
    return title[:15]


def read_visual_copy(product_name: str) -> dict[str, str | list[str]]:
    copy = PRODUCT_COPY.get(product_name)
    if copy:
        return copy

    md_path = PRODUCT_ROOT / "印花布品类" / "识别分组结果.md"
    if not md_path.exists():
        raise RuntimeError(f"{product_name} 缺少产品文案与识别描述。")

    lines = md_path.read_text(encoding="utf-8").splitlines()
    start = None
    for index, line in enumerate(lines):
        if line.strip() == f"- `{product_name}`":
            start = index
            break
    if start is None:
        raise RuntimeError(f"识别分组结果中未找到 {product_name}。")

    block: list[str] = []
    for line in lines[start + 1 :]:
        if line.startswith("- `") and line.strip() != f"- `{product_name}`":
            break
        if line.strip():
            block.append(line.strip())

    title_direction = ""
    visual_traits = ""
    for line in block:
        if line.startswith("- 标题方向："):
            title_direction = line.split("：", 1)[1].strip()
        if line.startswith("- 视觉特征："):
            visual_traits = line.split("：", 1)[1].strip()

    if not title_direction:
        title_direction = f"{product_name}棉布桌布"
    if not visual_traits:
        visual_traits = "印花棉布、统一米白流苏花边、轻薄自然垂感"

    return {
        "title": title_direction[:30],
        "guide_title": title_direction[:30],
        "title_direction": title_direction,
        "visual_traits": visual_traits,
        "style_tags": ["印花桌布", "棉布", "流苏边", "轻薄垂感"],
        "hero_title": title_direction[:8],
        "hero_subtitle": visual_traits[:24],
        "pattern_label": title_direction[:10],
        "scene_hint": "适合家用餐桌、茶几、边柜和生活化陈设",
    }


def read_context(product_name: str) -> ProductContext:
    product_dir = PRODUCT_ROOT / product_name
    if not product_dir.exists():
        raise RuntimeError(f"产品目录不存在：{product_dir}")

    workbook_path = next(product_dir.glob("*.xlsx"))
    images = sorted([path for path in product_dir.iterdir() if path.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}])
    texture_image = next((path for path in images if "_ref" in path.stem.lower()), None)
    if texture_image is None:
        texture_image = next((path for path in images if "_source" not in path.stem.lower()), None)
    if texture_image is None:
        raise RuntimeError(f"{product_dir} 缺少布样参考图。")

    local_edge = next((path for path in images if "edge" in path.stem.lower() or "花边" in path.stem), None)
    global_edge = PRODUCT_ROOT / "花边图" / "统一花边参考图.jpg"
    edge_image = local_edge if local_edge else global_edge
    if not edge_image.exists():
        raise RuntimeError(f"{product_dir} 缺少花边参考图。")

    wb = load_workbook(workbook_path)
    ws = wb["SKU信息登记"]
    sizes = [str(cell.value) for cell in ws[7][1:4] if cell.value]
    prices = [str(cell.value) for cell in ws[8][1:4] if cell.value]
    category = clean_text(ws["B4"].value, "桌布")
    material = clean_text(ws["B10"].value, "棉布")

    copy = read_visual_copy(product_name)
    title = clean_text(ws["B2"].value, str(copy["title"]))[:30]
    guide_title = clean_text(ws["B3"].value, str(copy["guide_title"]))[:30]
    style_raw = clean_text(ws["B6"].value, "、".join(copy["style_tags"]))
    style_tags = [item.strip() for item in style_raw.replace("，", "、").split("、") if item.strip()]

    return ProductContext(
        product_name=product_name,
        product_dir=product_dir,
        workbook_path=workbook_path,
        texture_image=texture_image,
        edge_image=edge_image,
        title=title,
        guide_title=guide_title,
        title_direction=str(copy["title_direction"]),
        house_style=str(copy["house_style"]),
        house_palette=str(copy["house_palette"]),
        visual_traits=str(copy["visual_traits"]),
        category=category,
        material=material,
        sizes=sizes,
        prices=prices,
        style_tags=style_tags,
        hero_title=str(copy["hero_title"]),
        hero_subtitle=str(copy["hero_subtitle"]),
        pattern_label=str(copy["pattern_label"]),
        scene_hint=str(copy["scene_hint"]),
    )


def build_prompt(
    ctx: ProductContext,
    *,
    aspect_label: str,
    scene_name: str,
    furniture_palette: str,
    decor_notes: str,
    composition: str,
    light: str,
    props: str,
    focus: str,
) -> str:
    return "\n".join(
        [
            "Use case: photorealistic-natural",
            f"Asset type: Taobao {aspect_label} tablecloth listing image",
            f"Primary request: generate one premium ecommerce lifestyle image for {ctx.title}",
            "Input images: Image 1 reference fabric pattern; Image 2 reference tassel edge",
            f"Scene/backdrop: {scene_name}",
            f"House style: {ctx.house_style}",
            f"House palette: {ctx.house_palette}",
            f"Furniture palette: {furniture_palette}",
            f"Decor notes: {decor_notes}",
            "Constraint: keep all images of this product in the same house style; vary only room function, table shape, angle, and props.",
            f"Subject: {ctx.visual_traits}; the product is a thin cotton printed tablecloth with an off-white tassel trim about 3 cm wide around the edges",
            "Style/medium: premium ecommerce photography, realistic home textile rendering",
            f"Composition/framing: {composition}",
            f"Lighting/mood: {light}",
            "Color palette: keep the product color accurate; warm neutral interior tones are allowed but must not shift the tablecloth colors",
            "Materials/textures: thin cotton print, clear woven textile texture, natural drape, no plastic sheen, no heavy thick fabric look",
            f"Constraints: keep the pattern accurate and readable; tassel trim must be visible and natural; {focus}; use props like {props}; no text; no watermark",
            "Avoid: collage, patchwork layout board, flat pasted fabric, floating cloth, excessive blur, deformed tassels, extra layers, synthetic sheen, thick upholstery look",
        ]
    )


def build_scene_specs(ctx: ProductContext) -> list[SceneSpec]:
    square_scenes = [
        ("sq01", "日常餐桌", "bright japandi dining room with a light oak table, linen curtains, and pale plaster wall", "warm oak table, beige spindle chairs, stoneware dishes", "simple wall art, dried branches, soft cream textiles, quiet modern home feel", "square composition, three-quarter angle, full tabletop visible, tassels visible on two sides", "soft side-window morning light", "plates, glassware, folded cloth napkin", "show the overall pattern clearly"),
        ("sq02", "庭院茶桌", "mediterranean courtyard table under a light parasol with stone walls and olive trees", "weathered wood table, white metal chair, terracotta planter accents", "ceramic jug, sun-faded cushions, layered greenery, outdoor hosting mood", "square composition, slightly elevated angle, one near corner visible, edge drape visible", "warm late-morning sun with gentle shade", "teapot, dessert plates, fruit dish", "show a fresh outdoor hosting atmosphere"),
        ("sq03", "阳台早餐", "paris balcony breakfast nook with a compact round table and iron bistro chair", "black iron chair, cream round table, warm beige floor tiles", "striped awning, small herb pot, glass vase, city-view lifestyle", "square composition, near eye-level angle, round tabletop centered, tassel edge flowing naturally", "soft morning glow", "coffee cup, toast plate, small vase", "show lightweight drape and everyday comfort"),
        ("sq04", "厨边陈列", "rustic kitchen island with sage cabinetry, open shelving, and stoneware", "sage cabinet, oak shelf, brass hooks, matte ceramic jars", "recipe books, candle, tray, branch leaves, lived-in kitchen styling", "square composition, frontal angle, tablecloth spread on a console or side cabinet with edge overhang", "warm afternoon light", "books, candle, tray, branch leaves", "show this is suitable beyond dining tables"),
        ("sq05", "花园桌角", "garden tea corner with a woven chair, terracotta pots, and blooming wildflowers", "wicker chair, terracotta planter, dark wood side table", "wildflower sprigs, linen napkin, rustic pottery, airy garden texture", "square composition, close-up angle, corner fold and tassels dominant, still clearly a tablecloth scene", "soft directional highlight", "ceramic cup, linen napkin, small plate", "show thin cotton texture and neat tassel finish"),
    ]
    portrait_scenes = [
        ("pt01", "餐桌主图", "minimal Scandinavian dining nook with pale oak furniture and stoneware", "pale oak table, cream upholstered chair, linen drape", "simple ceramic set, neutral wall art, airy lightness", "portrait composition, three-quarter angle, room depth visible behind the table", "clean side-window diffuse light", "plate set, glass cup, greenery", "make the tablecloth the main subject"),
        ("pt02", "待客氛围", "garden-facing dining scene with open French doors and a more layered home interior", "warm wood table, sage chair, light cane details", "flower vase, serving tray, open doorway, lived-in but refined home", "portrait composition, slightly wider frame, table and surrounding chair hints visible", "warm late-morning light", "tea set, serving tray, fruit bowl", "show richer decor without overpowering the product"),
        ("pt03", "圆桌茶点", "round tea table in a sunroom corner beside tall plants and a rattan seat", "round rattan-backed chair, cream tabletop, honey wood rim", "tea service, small blossoms, airy sunroom atmosphere", "portrait composition, gentle high angle, round table fully readable", "soft afternoon light", "tea cup, pastry plate, small flowers", "show a different table shape and scene rhythm"),
        ("pt04", "玄关边柜", "bright entry console styling scene with art prints and a brass tray", "walnut console, cream wall, black frame accents", "framed botanical art, brass tray, sculptural ornament, layered entryway", "portrait composition, frontal angle, cabinet styling emphasized", "soft indoor light with slight highlight separation", "frame, candle, stack of books, tray", "show multi-scene adaptability"),
        ("pt05", "近景垂感", "close scene focused on one corner of a dining table beside a window seat and soft bench", "cream bench, light oak table leg, beige floor", "minimal linen napkin, one cup, quiet close-up with breathable space", "portrait composition, close-up with one corner hanging naturally, tassels crisp", "controlled side light with texture emphasis", "plate rim, fork, cup", "show the lightweight cotton drape and edge workmanship"),
    ]
    detail_scenes = [
        ("dt01", "详情首屏", "warm japandi dining room hero scene with product fully styled", "light oak table, cream chair, linen curtain", "stoneware, pale art print, soft off-white accessories", "portrait composition, polished full-scene hero shot with clear negative space", "soft natural daylight", "plates, cup, foliage", "build first-screen purchase desire"),
        ("dt02", "花型说明", "tabletop scene emphasizing print readability and color accuracy in a clean modern kitchen", "white tile backdrop, charcoal stool, pale wood counter", "simple dishware, steel cup, restrained backdrop", "portrait composition, medium-close view across the tabletop", "soft window light", "simple dishware, no clutter", "focus on print and color explanation"),
        ("dt03", "材质流苏", "close edge scene showing tassels and textile weave on a dark wood bench", "dark walnut bench, warm brown floor", "one linen napkin, no extra decoration, texture-first close-up", "portrait composition, close framing on one hanging edge", "angled light with texture shadow", "minimal props only", "focus on texture, tassel trim, and thin cotton feel"),
        ("dt04", "多场景适配", "different home furniture setting such as console, tea table, or balcony table", "walnut console, rattan tea table, pale balcony chair", "books, tray, candle, mixed-home styling", "portrait composition, wider environment context", "warm indoor daylight", "books, tray, candle", "show scene adaptability without repeating main-image composition"),
        ("dt05", "尺寸选择", "clean listing-friendly dining scene with stable visual structure", "light oak table, ivory chair, calm wall", "minimal props, open negative space, size-selection clarity", "portrait composition, calm background and clear product silhouette", "balanced soft light", "minimal props", "leave visual rhythm suitable for a size selection page"),
    ]

    specs: list[SceneSpec] = [
        SceneSpec(
            scene_id="white01",
            bucket="white",
            aspect_label="白底图",
            size=(1536, 1536),
            headline=ctx.hero_title,
            subtitle=ctx.hero_subtitle,
            prompt="\n".join(
            [
                "Use case: product-mockup",
                "Asset type: Taobao white background product image",
                f"Primary request: create one photorealistic product image for {ctx.title}",
                "Input images: Image 1 reference fabric pattern; Image 2 reference tassel edge",
                "Scene/backdrop: pure clean white studio background",
                f"Subject: {ctx.visual_traits}; the tablecloth is neatly draped on a real solid dining table with a flat tabletop and sturdy legs that are mostly hidden, tassel trim about 3 cm wide visible on multiple sides, the hem hanging about 15 to 20 cm below the tabletop edge, corners relaxed and not stretched open, floor not visible or only a very small sliver at most",
                "Style/medium: clean ecommerce studio photography",
                "Composition/framing: square product-centered composition, clear margins, full product shape readable",
                "Lighting/mood: bright even studio light with soft grounding shadows only",
                "Color palette: keep the product colors accurate and neutral",
                "Materials/textures: thin cotton print, soft relaxed drape, crisp tassel strands, no plastic sheen, no heavy thick fabric feel, no board-like stiffness",
                "Constraints: realistic isolated catalog look; no text; no watermark; no decorative room scene; frame the table so the lower floor area is minimized",
                "Avoid: collage, floating fabric, clipped edges, overexposed whites, deformed tassels, hollow-looking table, warped tabletop, stretched corners, too much floor visible",
            ]
        ),
    )
    ]

    for scene_id, label, scene_name, composition, light, props, focus in square_scenes:
        specs.append(
            SceneSpec(
                scene_id=scene_id,
                bucket="square",
                aspect_label="1:1主图",
                size=(1536, 1536),
                headline=label,
                subtitle=focus,
                prompt=build_prompt(
                    ctx,
                    aspect_label="1:1 main image",
                    scene_name=scene_name,
                    furniture_palette=composition.split(",")[0] if "," in composition else composition,
                    decor_notes=focus,
                    composition=composition,
                    light=light,
                    props=props,
                    focus=focus,
                ),
            )
        )

    for scene_id, label, scene_name, composition, light, props, focus in portrait_scenes:
        specs.append(
            SceneSpec(
                scene_id=scene_id,
                bucket="portrait",
                aspect_label="3:4主图",
                size=(1536, 2048),
                headline=label,
                subtitle=focus,
                prompt=build_prompt(
                    ctx,
                    aspect_label="3:4 main image",
                    scene_name=scene_name,
                    furniture_palette=composition.split(",")[0] if "," in composition else composition,
                    decor_notes=focus,
                    composition=composition,
                    light=light,
                    props=props,
                    focus=focus,
                ),
            )
        )

    for scene_id, label, scene_name, composition, light, props, focus in detail_scenes:
        specs.append(
            SceneSpec(
                scene_id=scene_id,
                bucket="detail",
                aspect_label="详情页场景",
                size=(1536, 2048),
                headline=label,
                subtitle=focus,
                prompt=build_prompt(
                    ctx,
                    aspect_label="detail page image",
                    scene_name=scene_name,
                    furniture_palette=composition.split(",")[0] if "," in composition else composition,
                    decor_notes=focus,
                    composition=composition,
                    light=light,
                    props=props,
                    focus=focus,
                ),
            )
        )
    return specs


def get_cache_paths(product_name: str) -> dict[str, Path]:
    base = CACHE_ROOT / product_name
    raw_dir = base / "raw"
    refs_dir = base / "refs"
    manifest = base / "scene_manifest.json"
    prompts = base / "scene_prompts.md"
    review = base / "self_check.json"
    return {
        "base": base,
        "raw": raw_dir,
        "refs": refs_dir,
        "manifest": manifest,
        "prompts": prompts,
        "review": review,
    }


def write_prompt_book(ctx: ProductContext, specs: list[SceneSpec], prompts_path: Path) -> None:
    lines = [
        f"# {ctx.product_name} image-2 生图任务",
        "",
        f"- 标题：{ctx.title}",
        f"- 标题方向：{ctx.title_direction}",
        f"- 家居风格：{ctx.house_style}",
        f"- 家居配色：{ctx.house_palette}",
        f"- 视觉特征：{ctx.visual_traits}",
        "- 花边要求：统一米白流苏花边，视觉宽度约 3 厘米，排布参考 1 号品。",
        "- 布料要求：薄棉布，自然垂感，不做厚重面料。",
        "- 正式目录只保留最终成品，中间 raw 图留在缓存目录。",
        "",
    ]
    for spec in specs:
        lines.extend(
            [
                f"## {spec.scene_id} | {spec.aspect_label} | {spec.headline}",
                "",
                spec.prompt,
                "",
            ]
        )
    prompts_path.write_text("\n".join(lines), encoding="utf-8")


def prepare_product(product_name: str) -> dict[str, Path]:
    ctx = read_context(product_name)
    cache = get_cache_paths(product_name)
    ensure_dir(cache["base"])
    ensure_dir(cache["raw"])
    reset_dir(cache["refs"])
    shutil.copy2(ctx.texture_image, cache["refs"] / f"{product_name}_ref{ctx.texture_image.suffix.lower()}")
    shutil.copy2(ctx.edge_image, cache["refs"] / f"{product_name}_edge{ctx.edge_image.suffix.lower()}")

    specs = build_scene_specs(ctx)
    manifest_payload = {
        "product": asdict(ctx) | {
            "product_dir": str(ctx.product_dir),
            "workbook_path": str(ctx.workbook_path),
            "texture_image": str(ctx.texture_image),
            "edge_image": str(ctx.edge_image),
        },
        "cache_raw_dir": str(cache["raw"]),
        "scenes": [
            {
                "scene_id": spec.scene_id,
                "bucket": spec.bucket,
                "aspect_label": spec.aspect_label,
                "size": list(spec.size),
                "headline": spec.headline,
                "subtitle": spec.subtitle,
                "prompt": spec.prompt,
                "raw_path": str(cache["raw"] / f"{spec.scene_id}.png"),
            }
            for spec in specs
        ],
    }
    cache["manifest"].write_text(json.dumps(manifest_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_prompt_book(ctx, specs, cache["prompts"])
    return cache


def latest_generated_image() -> Path:
    if not GENERATED_IMAGE_ROOT.exists():
        raise RuntimeError("未找到订阅生图输出文件。")

    session_dirs = [path for path in GENERATED_IMAGE_ROOT.iterdir() if path.is_dir()]
    if session_dirs:
        latest_session = max(session_dirs, key=lambda path: path.stat().st_mtime)
        candidates = [path for path in latest_session.iterdir() if path.suffix.lower() in {".png", ".jpg", ".jpeg", ".webp"}]
        if candidates:
            return max(candidates, key=lambda path: path.stat().st_mtime)

    direct_files = [path for path in GENERATED_IMAGE_ROOT.iterdir() if path.is_file() and path.suffix.lower() in {".png", ".jpg", ".jpeg", ".webp"}]
    if direct_files:
        return max(direct_files, key=lambda path: path.stat().st_mtime)

    raise RuntimeError("未找到订阅生图输出文件。")


def capture_latest(product_name: str, scene_id: str, source: str | None = None) -> Path:
    cache = get_cache_paths(product_name)
    if not cache["manifest"].exists():
        raise RuntimeError("请先执行 prepare。")
    payload = json.loads(cache["manifest"].read_text(encoding="utf-8"))
    scene = next((item for item in payload["scenes"] if item["scene_id"] == scene_id), None)
    if scene is None:
        raise RuntimeError(f"未找到 scene_id：{scene_id}")
    src = Path(source) if source else latest_generated_image()
    ensure_dir(cache["raw"])
    dst = Path(scene["raw_path"])
    shutil.copy2(src, dst)
    return dst


def load_manifest(product_name: str) -> tuple[dict, dict[str, Path]]:
    cache = get_cache_paths(product_name)
    if not cache["manifest"].exists():
        raise RuntimeError("请先执行 prepare。")
    return json.loads(cache["manifest"].read_text(encoding="utf-8")), cache


def fit_image_to_canvas(src: Path, size: tuple[int, int]) -> Image.Image:
    return crop_cover(Image.open(src).convert("RGB"), size)


SQUARE_LAYOUT_COPY = {
    "sq01": {"eyebrow": "复古氛围", "title": "碎花落桌 餐厅更有温度", "body": "深底衬出木色纹理，日常一餐也显得从容。"},
    "sq02": {"eyebrow": "待客桌面", "title": "轻陈列一铺 画面便安静下来", "body": "花型收敛耐看，餐具与甜点更容易被衬得精致。"},
    "sq03": {"eyebrow": "晨光角落", "title": "小圆桌在晨光里更显松弛", "body": "布感轻，垂边柔，小空间仍留得出呼吸感。"},
    "sq04": {"eyebrow": "家居陈列", "title": "茶桌边柜 添一点旧时韵味", "body": "深色底更稳，摆进家里，比纯色布更见层次。"},
    "sq05": {"eyebrow": "近景质感", "title": "花型细密 流苏轻巧", "body": "近拍时布纹柔和清楚，边缘细节也显得秀气。"},
}

PORTRAIT_LAYOUT_COPY = {
    "pt01": {"eyebrow": "主图表达", "title": "深底黄白碎花", "body": "木色家具与这组小花最容易相互映衬。"},
    "pt02": {"eyebrow": "待客氛围", "title": "桌面安静下来 气质自然浮出来", "body": "聚餐、下午茶、甜点摆盘，画面愈发柔和。"},
    "pt03": {"eyebrow": "圆桌视角", "title": "小圆桌铺开 轻松节奏自然成形", "body": "布料轻薄，垂边自然，不会把空间压得发闷。"},
    "pt04": {"eyebrow": "空间延展", "title": "玄关与边柜 盛得下这份碎花气息", "body": "离开餐桌之后，它依旧适合家里的静物陈列。"},
    "pt05": {"eyebrow": "垂边细节", "title": "布纹柔和 流苏更显轻巧", "body": "边角近看更见质感，整体不会显得厚重。"},
}

DETAIL_LAYOUT_COPY = {
    "dt01": {
        "section": "01",
        "kicker": "首屏氛围",
        "title": "深底碎花铺上桌 画面先柔和了下来",
        "body": "黄白小花落在深底上，更容易衬出木色与奶油色空间。",
        "bullets": ["适合餐桌与下午茶场景", "第一眼便有复古气息", "不靠堆摆件也能成画面"],
    },
    "dt02": {
        "section": "02",
        "kicker": "花型配色",
        "title": "深底不显沉 反而更衬餐具与木色家具",
        "body": "黄白小花点缀其间，远看安静，近看也有细节可读。",
        "bullets": ["颜色稳，不容易显乱", "适合奶油风与原木风", "拍照时桌面层次更清楚"],
    },
    "dt03": {
        "section": "03",
        "kicker": "材质流苏",
        "title": "薄棉布垂落自然",
        "body": "不是厚重硬挺的质感，边缘垂下来会更柔和。",
        "bullets": ["布面显得轻盈", "流苏细节更完整", "近拍时不会笨重"],
    },
    "dt04": {
        "section": "04",
        "kicker": "多场景适配",
        "title": "一块布 把餐桌与陈列角落串成同一种气息",
        "body": "这类碎花不只属于用餐时刻，同样适合安静的家居布景。",
        "bullets": ["餐桌更有温度", "边柜陈列更完整", "民宿布景更相称"],
    },
    "dt05": {
        "section": "05",
        "kicker": "尺寸选择",
        "title": "按桌型挑尺寸 垂边比例自然更顺眼",
        "body": "从小桌到长桌，都能找到更舒服的落边尺度。",
        "bullets": ["60cm*60cm 适合小桌", "100cm*140cm 适合日常餐桌", "180cm*140cm 覆盖更完整"],
    },
}


def draw_text_block(
    draw: ImageDraw.ImageDraw,
    *,
    title: str,
    body: str,
    origin: tuple[int, int],
    title_font_size: int,
    body_font_size: int,
    title_fill: tuple[int, int, int, int],
    body_fill: str,
    title_width: int,
    body_width: int,
    line_gap: int = 8,
    max_height: int | None = None,
    min_title_font_size: int = 22,
    min_body_font_size: int = 16,
) -> int:
    x, y = origin
    title_width_px = int(title_font_size * title_width * 1.05)
    body_width_px = int(body_font_size * body_width * 1.05)
    if max_height is not None:
        while True:
            block_height, _, _ = measure_text_block(
                draw,
                title=title,
                body=body,
                title_font_size=title_font_size,
                body_font_size=body_font_size,
                title_width_px=title_width_px,
                body_width_px=body_width_px,
                line_gap=line_gap,
            )
            if block_height <= max_height:
                break
            shrunk = False
            if title_font_size > min_title_font_size:
                title_font_size -= 1
                title_width_px = int(title_font_size * title_width * 1.05)
                shrunk = True
            if body_font_size > min_body_font_size:
                body_font_size -= 1
                body_width_px = int(body_font_size * body_width * 1.05)
                shrunk = True
            if not shrunk:
                break

    title_font = load_font(title_font_size)
    for line in wrap_text_pixels(draw, title, title_font, title_width_px):
        draw.text((x, y), line, font=title_font, fill=title_fill)
        y += title_font.size + line_gap
    y += 8
    body_font = load_font(body_font_size)
    for line in wrap_text_pixels(draw, body, body_font, body_width_px):
        draw.text((x, y), line, font=body_font, fill=body_fill)
        y += body_font.size + line_gap
    return y


def draw_tag(
    draw: ImageDraw.ImageDraw,
    *,
    xy: tuple[int, int],
    text: str,
    fill: tuple[int, int, int, int],
    text_fill: tuple[int, int, int, int],
    font_size: int = 22,
    padding_x: int = 20,
    padding_y: int = 10,
    radius: int = 18,
) -> tuple[int, int, int, int]:
    font = load_font(font_size)
    bbox = draw.textbbox((0, 0), text, font=font)
    width = bbox[2] - bbox[0]
    height = bbox[3] - bbox[1]
    x, y = xy
    rect = (x, y, x + width + padding_x * 2, y + height + padding_y * 2)
    draw.rounded_rectangle(rect, radius=radius, fill=fill)
    draw.text((x + padding_x, y + padding_y - 2), text, font=font, fill=text_fill)
    return rect


def draw_detail_bullets(
    draw: ImageDraw.ImageDraw,
    items: list[str],
    start_xy: tuple[int, int],
    width: int,
    *,
    font_size: int = 24,
    step: int = 66,
    text_fill: str = "#4E5661",
    bullet_fill: tuple[int, int, int, int] = (150, 116, 79, 255),
) -> None:
    x, y = start_xy
    bullet_font = load_font(font_size)
    for item in items:
        draw.ellipse((x, y + 12, x + 10, y + 22), fill=bullet_fill)
        draw_multiline(draw, item, (x + 24, y), bullet_font, text_fill, line_gap=6, max_chars=width)
        y += step


def render_square_layout(img: Image.Image, scene_id: str) -> Image.Image:
    copy = SQUARE_LAYOUT_COPY[scene_id]
    canvas = img.convert("RGBA")
    overlay = Image.new("RGBA", canvas.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)

    if scene_id == "sq01":
        draw.rounded_rectangle((36, 40, 410, 270), radius=34, fill=(245, 240, 233, 212))
        draw.rounded_rectangle((54, 118, 66, 250), radius=6, fill=(180, 143, 104, 228))
        draw_tag(draw, xy=(62, 62), text=copy["eyebrow"], fill=(90, 74, 61, 235), text_fill=(252, 247, 240, 255))
        draw_text_block(draw, title=copy["title"], body=copy["body"], origin=(64, 128), title_font_size=46, body_font_size=23, title_fill=(77, 57, 42, 255), body_fill="#7D6A59", title_width=8, body_width=12, max_height=112)
    elif scene_id == "sq02":
        draw.rounded_rectangle((420, 54, 744, 372), radius=34, fill=(39, 34, 31, 176))
        draw.rectangle((446, 146, 456, 330), fill=(222, 192, 153, 230))
        draw_tag(draw, xy=(476, 82), text=copy["eyebrow"], fill=(233, 219, 194, 245), text_fill=(96, 71, 45, 255))
        draw_text_block(draw, title=copy["title"], body=copy["body"], origin=(466, 148), title_font_size=30, body_font_size=17, title_fill=(249, 239, 228, 255), body_fill="#E7D8C8", title_width=8, body_width=13, max_height=190)
    elif scene_id == "sq03":
        draw.rounded_rectangle((54, 510, 364, 750), radius=28, fill=(249, 244, 237, 224))
        draw.rectangle((78, 556, 298, 558), fill=(180, 140, 94, 235))
        draw_tag(draw, xy=(78, 578), text=copy["eyebrow"], fill=(181, 137, 94, 230), text_fill=(255, 250, 246, 255))
        draw_text_block(draw, title=copy["title"], body=copy["body"], origin=(80, 632), title_font_size=30, body_font_size=18, title_fill=(95, 66, 44, 255), body_fill="#6F5B48", title_width=8, body_width=11, max_height=102)
    elif scene_id == "sq04":
        draw.rounded_rectangle((42, 52, 740, 242), radius=32, fill=(244, 238, 229, 210))
        draw.rectangle((70, 82, 82, 214), fill=(133, 111, 88, 230))
        draw_tag(draw, xy=(98, 82), text=copy["eyebrow"], fill=(118, 101, 81, 235), text_fill=(251, 245, 237, 255))
        draw_text_block(draw, title=copy["title"], body=copy["body"], origin=(100, 144), title_font_size=34, body_font_size=21, title_fill=(79, 58, 45, 255), body_fill="#746857", title_width=12, body_width=19, max_height=76)
    else:
        draw.rounded_rectangle((58, 72, 320, 302), radius=34, fill=(248, 244, 239, 214))
        draw_tag(draw, xy=(84, 98), text=copy["eyebrow"], fill=(88, 73, 60, 236), text_fill=(248, 243, 236, 255))
        draw_text_block(draw, title=copy["title"], body=copy["body"], origin=(84, 158), title_font_size=32, body_font_size=20, title_fill=(77, 57, 42, 255), body_fill="#736655", title_width=7, body_width=9, max_height=116)

    return Image.alpha_composite(canvas, overlay).convert("RGB")


def render_portrait_layout(img: Image.Image, scene_id: str) -> Image.Image:
    copy = PORTRAIT_LAYOUT_COPY[scene_id]
    canvas = img.convert("RGBA")
    overlay = Image.new("RGBA", canvas.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)

    if scene_id == "pt01":
        draw.rounded_rectangle((52, 60, 560, 312), radius=34, fill=(248, 243, 236, 210))
        draw_tag(draw, xy=(82, 88), text=copy["eyebrow"], fill=(87, 73, 60, 235), text_fill=(251, 245, 236, 255))
        draw_text_block(draw, title=copy["title"], body=copy["body"], origin=(84, 150), title_font_size=62, body_font_size=28, title_fill=(74, 55, 43, 255), body_fill="#796655", title_width=8, body_width=14, max_height=132)
    elif scene_id == "pt02":
        draw.rounded_rectangle((58, 808, 520, 1148), radius=34, fill=(248, 242, 234, 220))
        draw.rectangle((84, 898, 96, 1092), fill=(185, 149, 110, 230))
        draw_tag(draw, xy=(108, 866), text=copy["eyebrow"], fill=(234, 220, 191, 245), text_fill=(104, 78, 49, 255))
        draw_text_block(draw, title=copy["title"], body=copy["body"], origin=(110, 914), title_font_size=34, body_font_size=20, title_fill=(88, 66, 47, 255), body_fill="#7B6756", title_width=8, body_width=12, max_height=210)
    elif scene_id == "pt03":
        draw.rounded_rectangle((470, 70, 842, 342), radius=32, fill=(246, 239, 230, 220))
        draw.rectangle((502, 100, 808, 104), fill=(185, 141, 96, 235))
        draw_tag(draw, xy=(502, 124), text=copy["eyebrow"], fill=(185, 141, 96, 235), text_fill=(255, 248, 241, 255))
        draw_text_block(draw, title=copy["title"], body=copy["body"], origin=(504, 186), title_font_size=36, body_font_size=21, title_fill=(93, 67, 45, 255), body_fill="#7B6753", title_width=7, body_width=9, max_height=118)
    elif scene_id == "pt04":
        draw.rounded_rectangle((64, 72, 422, 352), radius=34, fill=(244, 239, 232, 216))
        draw.rounded_rectangle((92, 108, 264, 160), radius=24, fill=(112, 93, 74, 236))
        draw.text((120, 120), copy["eyebrow"], font=load_font(24), fill=(250, 244, 236, 255))
        draw_text_block(draw, title=copy["title"], body=copy["body"], origin=(96, 182), title_font_size=34, body_font_size=19, title_fill=(77, 57, 44, 255), body_fill="#6F655A", title_width=8, body_width=12, max_height=146)
    else:
        draw.rounded_rectangle((52, 940, 846, 1142), radius=30, fill=(248, 244, 238, 214))
        draw_tag(draw, xy=(86, 972), text=copy["eyebrow"], fill=(88, 73, 61, 236), text_fill=(248, 244, 236, 255))
        draw_text_block(draw, title=copy["title"], body=copy["body"], origin=(250, 972), title_font_size=36, body_font_size=20, title_fill=(78, 58, 44, 255), body_fill="#776856", title_width=10, body_width=16, max_height=140)

    return Image.alpha_composite(canvas, overlay).convert("RGB")


def render_detail_layout(src: Path, scene_id: str, out_path: Path) -> None:
    copy = DETAIL_LAYOUT_COPY[scene_id]
    hero = fit_image_to_canvas(src, (750, 1200))
    canvas = hero.convert("RGBA")
    overlay = Image.new("RGBA", canvas.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)

    if scene_id == "dt01":
        draw.rounded_rectangle((34, 824, 716, 1178), radius=38, fill=(247, 241, 233, 210))
        draw.text((62, 848), copy["section"], font=load_font(28), fill=(176, 134, 93, 255))
        draw.text((116, 852), copy["kicker"], font=load_font(24), fill=(145, 112, 82, 255))
        end_y = draw_text_block(draw, title=copy["title"], body=copy["body"], origin=(60, 900), title_font_size=35, body_font_size=20, title_fill=(77, 57, 42, 255), body_fill="#776555", title_width=12, body_width=21, max_height=120)
        draw_detail_bullets(draw, copy["bullets"], (60, end_y), 18, font_size=20, step=48)
    elif scene_id == "dt02":
        draw.rounded_rectangle((44, 790, 706, 1174), radius=34, fill=(34, 30, 28, 176))
        draw.text((72, 818), copy["section"], font=load_font(28), fill=(235, 215, 187, 255))
        draw.text((126, 822), copy["kicker"], font=load_font(24), fill=(230, 211, 190, 255))
        end_y = draw_text_block(draw, title=copy["title"], body=copy["body"], origin=(70, 872), title_font_size=37, body_font_size=21, title_fill=(251, 244, 236, 255), body_fill="#F2E7D7", title_width=12, body_width=20, max_height=132)
        draw_detail_bullets(draw, copy["bullets"], (72, end_y), 16, font_size=21, step=50, text_fill="#F0E3D0", bullet_fill=(236, 210, 173, 255))
    elif scene_id == "dt03":
        draw.rounded_rectangle((44, 60, 362, 516), radius=34, fill=(245, 238, 230, 214))
        draw.text((74, 86), copy["section"], font=load_font(28), fill=(176, 134, 93, 255))
        draw.text((128, 90), copy["kicker"], font=load_font(24), fill=(145, 111, 80, 255))
        end_y = draw_text_block(draw, title=copy["title"], body=copy["body"], origin=(74, 142), title_font_size=32, body_font_size=20, title_fill=(77, 57, 42, 255), body_fill="#776758", title_width=7, body_width=9, max_height=138)
        draw_detail_bullets(draw, copy["bullets"], (76, end_y + 10), 8)
    elif scene_id == "dt04":
        draw.rounded_rectangle((44, 66, 496, 430), radius=34, fill=(247, 243, 236, 214))
        draw.text((74, 94), copy["section"], font=load_font(28), fill=(176, 134, 93, 255))
        draw.text((128, 98), copy["kicker"], font=load_font(24), fill=(145, 111, 82, 255))
        end_y = draw_text_block(draw, title=copy["title"], body=copy["body"], origin=(74, 150), title_font_size=29, body_font_size=19, title_fill=(79, 58, 43, 255), body_fill="#756857", title_width=10, body_width=13, max_height=112)
        draw_detail_bullets(draw, copy["bullets"], (76, end_y + 6), 10, font_size=21, step=52)
    else:
        draw.rounded_rectangle((34, 44, 716, 392), radius=36, fill=(248, 243, 236, 208))
        draw.text((62, 64), copy["section"], font=load_font(30), fill=(176, 134, 93, 255))
        draw.text((120, 68), copy["kicker"], font=load_font(26), fill=(145, 111, 82, 255))
        end_y = draw_text_block(draw, title=copy["title"], body=copy["body"], origin=(60, 120), title_font_size=38, body_font_size=21, title_fill=(79, 58, 43, 255), body_fill="#756857", title_width=13, body_width=20, max_height=132)
        draw_detail_bullets(draw, copy["bullets"], (60, end_y + 6), 16, font_size=20, step=50)

    Image.alpha_composite(canvas, overlay).convert("RGB").save(out_path, quality=95)


def average_hash(path: Path, size: int = 8) -> str:
    img = Image.open(path).convert("L").resize((size, size), Image.Resampling.LANCZOS)
    pixels = list(img.getdata())
    avg = sum(pixels) / len(pixels)
    bits = "".join("1" if pixel >= avg else "0" for pixel in pixels)
    return f"{int(bits, 2):016x}"


def hamming_distance(a: str, b: str) -> int:
    return bin(int(a, 16) ^ int(b, 16)).count("1")


def write_self_check(product_name: str, payload: dict, final_paths: list[Path], review_path: Path) -> None:
    hashes = {path.name: average_hash(path) for path in final_paths}
    duplicate_pairs = []
    names = list(hashes)
    for i, name_a in enumerate(names):
        for name_b in names[i + 1 :]:
            dist = hamming_distance(hashes[name_a], hashes[name_b])
            if dist <= 6:
                duplicate_pairs.append({"a": name_a, "b": name_b, "distance": dist})
    report = {
        "product": product_name,
        "generated_scene_count": len(payload["scenes"]),
        "final_asset_count": len(final_paths),
        "checks": {
            "raw_scene_count_ready": all(Path(scene["raw_path"]).exists() for scene in payload["scenes"]),
            "formal_directories_final_only": True,
            "potential_duplicate_pairs": duplicate_pairs,
            "manual_review": {
                "product_accuracy": "pending",
                "pattern_color_accuracy": "pending",
                "dot_or_floral_scale_accuracy": "pending",
                "tassel_and_edge_accuracy": "pending",
                "fabric_not_too_thick": "pending",
                "copy_not_overlapping": "pending",
                "layout_not_template_like": "pending",
                "scene_diversity_clear": "pending",
                "detail_pages_not_repeating_main_images": "pending",
            },
        },
    }
    review_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")


def update_workbook(ctx: ProductContext) -> None:
    wb = load_workbook(ctx.workbook_path)
    ws = wb["SKU信息登记"]
    ws["B2"] = build_listing_title(ctx)
    ws["B3"] = build_guide_title(ctx)
    ws["B6"] = "、".join(ctx.style_tags)
    if str(ws["B27"].value or "").strip() == "3厘米":
        ws["B27"] = "无"
    wb.save(ctx.workbook_path)


def publish_product(product_name: str) -> list[Path]:
    payload, cache = load_manifest(product_name)
    ctx = read_context(product_name)
    scenes = payload["scenes"]
    missing = [scene["scene_id"] for scene in scenes if not Path(scene["raw_path"]).exists()]
    if missing:
        raise RuntimeError(f"以下 raw 图缺失，无法正式出图：{', '.join(missing)}")

    white_dir = ctx.product_dir / "白底图"
    square_dir = ctx.product_dir / "1比1主图"
    portrait_dir = ctx.product_dir / "3比4主图"
    detail_dir = ctx.product_dir / "详情页"
    for path in [white_dir, square_dir, portrait_dir, detail_dir]:
        reset_dir(path)

    final_paths: list[Path] = []

    white_scene = next(scene for scene in scenes if scene["bucket"] == "white")
    white_out = white_dir / f"{product_name}白底图1.png"
    shutil.copy2(white_scene["raw_path"], white_out)
    final_paths.append(white_out)

    main_index = 1
    square_scenes = [scene for scene in scenes if scene["bucket"] == "square"]
    for scene in square_scenes:
        rendered = render_square_layout(fit_image_to_canvas(Path(scene["raw_path"]), (800, 800)), scene["scene_id"])
        out = square_dir / f"{product_name}主图{main_index}.jpg"
        rendered.save(out, quality=95)
        final_paths.append(out)
        main_index += 1

    portrait_scenes = [scene for scene in scenes if scene["bucket"] == "portrait"]
    for scene in portrait_scenes:
        rendered = render_portrait_layout(fit_image_to_canvas(Path(scene["raw_path"]), (900, 1200)), scene["scene_id"])
        out = portrait_dir / f"{product_name}主图{main_index}.jpg"
        rendered.save(out, quality=95)
        final_paths.append(out)
        main_index += 1
    detail_scenes = [scene for scene in scenes if scene["bucket"] == "detail"]
    for index, scene in enumerate(detail_scenes, start=1):
        out = detail_dir / f"{product_name}详情图{index}.jpg"
        render_detail_layout(Path(scene["raw_path"]), scene["scene_id"], out)
        final_paths.append(out)

    update_workbook(ctx)
    write_self_check(product_name, payload, final_paths, cache["review"])
    return final_paths


def print_status(product_name: str) -> None:
    payload, _ = load_manifest(product_name)
    ready = []
    missing = []
    for scene in payload["scenes"]:
        raw_path = Path(scene["raw_path"])
        (ready if raw_path.exists() else missing).append(scene["scene_id"])
    print("READY:", ", ".join(ready) if ready else "NONE")
    print("MISSING:", ", ".join(missing) if missing else "NONE")


def main() -> None:
    parser = argparse.ArgumentParser(description="桌布 image-2 真实场景工作流")
    subparsers = parser.add_subparsers(dest="command", required=True)

    prepare_parser = subparsers.add_parser("prepare", help="生成 raw 缓存目录、scene manifest 和 prompts")
    prepare_parser.add_argument("--product", required=True, help="产品目录名，例如 2号品")

    capture_parser = subparsers.add_parser("capture-latest", help="把最近一张订阅生图抓到指定 scene_id")
    capture_parser.add_argument("--product", required=True)
    capture_parser.add_argument("--scene-id", required=True)
    capture_parser.add_argument("--source", help="可选：指定原图路径，不指定则抓最近一张订阅生图")

    publish_parser = subparsers.add_parser("publish", help="把 raw 图排版为正式目录成品")
    publish_parser.add_argument("--product", required=True)

    status_parser = subparsers.add_parser("status", help="查看 scene raw 图是否齐全")
    status_parser.add_argument("--product", required=True)

    args = parser.parse_args()

    if args.command == "prepare":
        paths = prepare_product(args.product)
        print(paths["manifest"])
        print(paths["prompts"])
        print(paths["raw"])
        return
    if args.command == "capture-latest":
        print(capture_latest(args.product, args.scene_id, args.source))
        return
    if args.command == "publish":
        for path in publish_product(args.product):
            print(path)
        return
    if args.command == "status":
        print_status(args.product)
        return


if __name__ == "__main__":
    main()
