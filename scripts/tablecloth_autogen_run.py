from __future__ import annotations

import argparse
import math
import textwrap
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageColor, ImageDraw, ImageFilter, ImageFont


ROOT = Path(__file__).resolve().parents[1]
WORKFLOW_ROOT = ROOT / "桌布自动生图工作流"
PRODUCT_ROOT = WORKFLOW_ROOT / "产品素材"
FONT_PATH = Path(r"C:\Windows\Fonts\msyh.ttc")


@dataclass
class ProductContext:
    product_dir: Path
    workbook_path: Path
    texture_image: Path
    edge_image: Path
    title: str
    guide_title: str
    category: str
    material: str
    sizes: list[str]
    prices: list[str]
    style_tags: list[str]
    hero_title: str
    hero_subtitle: str
    pattern_label: str
    fringe_label: str
    size_label: str
    detail_material: str
    detail_fringe: str
    detail_scene: str


PRODUCT_COPY = {
    "1号品": {
        "title": "北欧米白深灰蓝波点流苏桌布 亚麻感餐桌茶几盖布",
        "guide_title": "米白蓝灰波点流苏桌布清爽百搭款",
        "style_tags": ["北欧简约", "米白蓝灰点", "流苏花边", "低饱和百搭", "亚麻感"],
        "hero_title": "米白蓝灰点",
        "hero_subtitle": "颜色更柔和，适合餐桌和茶几",
        "pattern_label": "米白蓝灰点",
        "fringe_label": "流苏边更出片",
        "size_label": "三档尺寸",
        "detail_material": "亚麻感肌理清楚，不会只有远看好看。",
        "detail_fringe": "流苏边是这轮素材里最明显的识别点。",
        "detail_scene": "更适合家用、茶几、拍照布景和民宿陈列。",
    },
    "2号品": {
        "title": "复古深底碎花流苏棉布桌布",
        "guide_title": "深底黄白碎花流苏桌布复古百搭款",
        "style_tags": ["复古田园", "深底碎花", "米白流苏", "棉布印花", "轻薄垂感"],
        "hero_title": "深底黄白碎花",
        "hero_subtitle": "复古田园感更明显，木色餐桌更好搭",
        "pattern_label": "深底碎花更耐看",
        "fringe_label": "3cm流苏花边",
        "size_label": "常用尺寸",
        "detail_material": "棉布印花偏轻薄，布面柔和，近看有细密织纹。",
        "detail_fringe": "米白流苏花边宽约3厘米，排布参考1号品布局。",
        "detail_scene": "更适合餐桌、早餐角、边柜和复古家居陈列。",
    },
}


def load_font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    if FONT_PATH.exists():
        return ImageFont.truetype(str(FONT_PATH), size=size)
    return ImageFont.load_default()


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def ensure_output_dir(product_dir: Path, preferred_name: str, legacy_name: str) -> Path:
    preferred_path = product_dir / preferred_name
    legacy_path = product_dir / legacy_name
    if legacy_path.exists() and not preferred_path.exists():
        legacy_path.rename(preferred_path)
    ensure_dir(preferred_path)
    return preferred_path


def wrap_text(text: str, width: int) -> list[str]:
    return textwrap.wrap(text, width=width, break_long_words=False, break_on_hyphens=False)


def draw_multiline(
    draw: ImageDraw.ImageDraw,
    text: str,
    xy: tuple[int, int],
    font: ImageFont.ImageFont,
    fill: str,
    line_gap: int = 8,
    max_chars: int = 14,
) -> int:
    x, y = xy
    lines = wrap_text(text, max_chars)
    for line in lines:
        draw.text((x, y), line, font=font, fill=fill)
        y += font.size + line_gap
    return y


def crop_cover(source: Image.Image, size: tuple[int, int], focus_bottom: bool = False) -> Image.Image:
    target_w, target_h = size
    src_ratio = source.width / source.height
    dst_ratio = target_w / target_h
    if src_ratio > dst_ratio:
        new_h = target_h
        new_w = int(new_h * src_ratio)
    else:
        new_w = target_w
        new_h = int(new_w / src_ratio)
    resized = source.resize((new_w, new_h), Image.Resampling.LANCZOS)
    left = max((new_w - target_w) // 2, 0)
    if focus_bottom:
        top = max(new_h - target_h - 40, 0)
    else:
        top = max((new_h - target_h) // 2, 0)
    return resized.crop((left, top, left + target_w, top + target_h))


def clean_edge_source(source: Image.Image) -> Image.Image:
    # The source close-up includes a phone shadow in the bottom-right corner.
    # Trim that area once up front so all derived layouts stay cleaner.
    right = int(source.width * 0.8)
    bottom = int(source.height * 0.9)
    return source.crop((0, 0, right, bottom))


def add_shadow(card: Image.Image, blur: int = 20, offset: tuple[int, int] = (0, 18)) -> Image.Image:
    shadow = Image.new("RGBA", card.size, (0, 0, 0, 0))
    mask = Image.new("L", card.size, 0)
    mask_draw = ImageDraw.Draw(mask)
    mask_draw.rounded_rectangle((18, 18, card.width - 18, card.height - 18), radius=34, fill=170)
    shadow.paste((0, 0, 0, 110), (0, 0), mask)
    shadow = shadow.filter(ImageFilter.GaussianBlur(blur))
    base = Image.new("RGBA", (card.width + abs(offset[0]) + 40, card.height + abs(offset[1]) + 40), (0, 0, 0, 0))
    base.alpha_composite(shadow, (20 + max(offset[0], 0), 20 + max(offset[1], 0)))
    base.alpha_composite(card, (20, 20))
    return base


def read_context(product_name: str | None = None) -> ProductContext:
    if product_name:
        product_dir = PRODUCT_ROOT / product_name
    else:
        product_dir = next(path for path in PRODUCT_ROOT.iterdir() if path.is_dir())
    workbook_path = next(product_dir.glob("*.xlsx"))
    images = sorted([path for path in product_dir.iterdir() if path.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}])
    texture_image = next((path for path in images if "_ref" in path.stem.lower()), None)
    if texture_image is None:
        texture_image = next((path for path in images if "_source" not in path.stem.lower()), None)
    if texture_image is None:
        raise RuntimeError(f"{product_dir} 缺少可用的布面参考图。")

    local_edge = next((path for path in images if "edge" in path.stem.lower() or "花边" in path.stem), None)
    global_edge = PRODUCT_ROOT / "花边图" / "统一花边参考图.jpg"
    edge_image = local_edge if local_edge else global_edge
    if not edge_image.exists():
        raise RuntimeError(f"{product_dir} 缺少花边参考图。")

    wb = load_workbook(workbook_path)
    ws = wb["SKU信息登记"]
    category = str(ws["B4"].value or "桌布")
    material = str(ws["B10"].value or "棉布")
    sizes = [str(cell.value) for cell in ws[7][1:4] if cell.value]
    prices = [str(cell.value) for cell in ws[8][1:4] if cell.value]
    copy = PRODUCT_COPY.get(product_dir.name, PRODUCT_COPY["1号品"])

    title = str(ws["B2"].value or copy["title"])
    guide_title = str(ws["B3"].value or copy["guide_title"])
    style_raw = str(ws["B6"].value or "、".join(copy["style_tags"]))
    style_tags = [item.strip() for item in style_raw.replace("，", "、").split("、") if item.strip()]

    return ProductContext(
        product_dir=product_dir,
        workbook_path=workbook_path,
        texture_image=texture_image,
        edge_image=edge_image,
        title=title,
        guide_title=guide_title,
        category=category,
        material=material,
        sizes=sizes,
        prices=prices,
        style_tags=style_tags,
        hero_title=copy["hero_title"],
        hero_subtitle=copy["hero_subtitle"],
        pattern_label=copy["pattern_label"],
        fringe_label=copy["fringe_label"],
        size_label=copy["size_label"],
        detail_material=copy["detail_material"],
        detail_fringe=copy["detail_fringe"],
        detail_scene=copy["detail_scene"],
    )


def make_white_bg(texture: Image.Image, edge: Image.Image, output_path: Path) -> None:
    canvas = Image.new("RGB", (1200, 1200), "#FFFFFF")
    draw = ImageDraw.Draw(canvas)
    draw.rounded_rectangle((80, 120, 1120, 1080), radius=48, fill="#FCFCFB")

    card = Image.new("RGBA", (900, 900), (255, 255, 255, 0))
    swatch = crop_cover(texture, (760, 540))
    swatch = swatch.convert("RGBA")
    swatch_mask = Image.new("L", swatch.size, 255)
    mask_draw = ImageDraw.Draw(swatch_mask)
    mask_draw.rounded_rectangle((0, 0, swatch.width, swatch.height), radius=26, fill=255)
    swatch.putalpha(swatch_mask)

    edge_band = crop_cover(edge, (760, 210), focus_bottom=True).convert("RGBA")
    edge_band.putalpha(255)

    card.alpha_composite(swatch, (70, 90))
    card.alpha_composite(edge_band, (70, 600))
    shadowed = add_shadow(card, blur=24, offset=(0, 24))
    canvas.paste(shadowed.convert("RGB"), (120, 120), shadowed)

    title_font = load_font(36)
    sub_font = load_font(24)
    draw.text((110, 54), "白底图最终版", font=title_font, fill="#1F2430")
    draw.text((110, 98), "桌体要实、布感要软、边角放松、离地不贴地", font=sub_font, fill="#6B7280")
    canvas.save(output_path, quality=95)


def render_square_series(ctx: ProductContext, texture: Image.Image, edge: Image.Image, output_dir: Path) -> list[Path]:
    outputs: list[Path] = []
    bg_colors = ["#F7F3ED", "#F5F5F4", "#EEE8E0", "#F2EFEA", "#F7F6F2"]
    accents = ["#1F2937", "#4B5563", "#374151", "#6B7280", "#111827"]
    headlines = [
        (ctx.hero_title, ctx.hero_subtitle),
        (ctx.fringe_label, "边缘细节比普通平边更完整"),
        ("布面纹理", ctx.detail_material),
        (ctx.size_label, "小桌到餐桌都能覆盖"),
        ("日常更好搭", ctx.detail_scene),
    ]

    for idx, (headline, body) in enumerate(headlines, start=1):
        canvas = Image.new("RGB", (800, 800), bg_colors[idx - 1])
        draw = ImageDraw.Draw(canvas)
        draw.rounded_rectangle((36, 36, 764, 764), radius=34, fill="#FFFDFC")

        hero = crop_cover(edge if idx % 2 else texture, (430, 540), focus_bottom=True)
        hero = hero.rotate(-3 if idx in {1, 4} else 3, expand=True, fillcolor=ImageColor.getrgb("#FFFDFC"))
        canvas.paste(hero, (322, 136))

        draw.text((72, 110), headline, font=load_font(58), fill=accents[idx - 1])
        draw_multiline(draw, body, (76, 190), load_font(26), "#6B7280", max_chars=12)
        draw.rounded_rectangle((74, 654, 286, 714), radius=22, fill="#111827")
        draw.text((102, 668), f"1:1 主图 {idx:02d}", font=load_font(24), fill="#FFFFFF")
        draw.line((76, 612, 284, 612), fill="#D6D3D1", width=2)
        draw.text((76, 574), "试跑版", font=load_font(22), fill="#9CA3AF")

        output_path = output_dir / f"02_1比1主图_{idx:02d}.jpg"
        canvas.save(output_path, quality=95)
        outputs.append(output_path)

    return outputs


def render_portrait_series(ctx: ProductContext, texture: Image.Image, edge: Image.Image, output_dir: Path) -> list[Path]:
    outputs: list[Path] = []
    page_titles = [
        (ctx.hero_title, ctx.hero_subtitle),
        ("花边细节", ctx.detail_fringe),
        ("家用陈列", ctx.detail_scene),
        ("尺寸选择", " / ".join(ctx.sizes)),
        ("近拍细节", ctx.detail_material),
    ]
    for idx, (title, sub) in enumerate(page_titles, start=1):
        canvas = Image.new("RGB", (900, 1200), "#F6F2EC")
        draw = ImageDraw.Draw(canvas)
        draw.rounded_rectangle((50, 52, 850, 1148), radius=42, fill="#FFFDF9")

        hero = crop_cover(edge if idx in {2, 5} else texture, (760, 620), focus_bottom=idx in {2, 5})
        canvas.paste(hero, (70, 120))
        draw.text((88, 792), title, font=load_font(64), fill="#1F2937")
        draw_multiline(draw, sub, (88, 878), load_font(30), "#6B7280", max_chars=14)

        if idx == 4:
            y = 960
            for size, price in zip(ctx.sizes, ctx.prices):
                draw.rounded_rectangle((88, y, 812, y + 78), radius=20, fill="#F3EFE8")
                draw.text((118, y + 18), size, font=load_font(28), fill="#374151")
                draw.text((620, y + 18), price, font=load_font(28), fill="#111827")
                y += 92
        else:
            draw.rounded_rectangle((88, 988, 320, 1050), radius=24, fill="#111827")
            draw.text((118, 1006), f"3:4 主图 {idx:02d}", font=load_font(24), fill="#FFFFFF")

        output_path = output_dir / f"03_3比4主图_{idx:02d}.jpg"
        canvas.save(output_path, quality=95)
        outputs.append(output_path)
    return outputs


def render_selling_point(ctx: ProductContext, texture: Image.Image, edge: Image.Image, output_path: Path) -> None:
    canvas = Image.new("RGB", (750, 1200), "#F7F3EE")
    draw = ImageDraw.Draw(canvas)
    draw.text((54, 66), "这款桌布先看 4 个点", font=load_font(46), fill="#1F2937")
    draw.text((54, 122), "成品卖点图，重点放真实可见细节", font=load_font(24), fill="#6B7280")

    cards = [
        ("花型耐看", ctx.hero_subtitle),
        ("花边加分", ctx.detail_fringe),
        ("尺寸完整", "从小桌到长桌，常用规格都能覆盖。"),
        ("近拍有料", ctx.detail_material),
    ]
    for idx, (title, body) in enumerate(cards):
        top = 208 + idx * 224
        draw.rounded_rectangle((44, top, 706, top + 184), radius=28, fill="#FFFDFC")
        thumb = crop_cover(edge if idx % 2 else texture, (176, 144), focus_bottom=bool(idx % 2))
        canvas.paste(thumb, (60, top + 20))
        draw.text((262, top + 28), title, font=load_font(34), fill="#111827")
        draw_multiline(draw, body, (262, top + 84), load_font(24), "#6B7280", max_chars=16)

    canvas.save(output_path, quality=95)


def render_detail_pages(ctx: ProductContext, texture: Image.Image, edge: Image.Image, output_dir: Path) -> tuple[Path, list[Path]]:
    page_w, page_h = 750, 6000
    canvas = Image.new("RGB", (page_w, page_h), "#F8F5F0")
    draw = ImageDraw.Draw(canvas)

    sections = [
        ("01 主题封面", f"{ctx.hero_title} + {ctx.fringe_label}，先建立产品印象。"),
        ("02 布面纹理", f"材质：{ctx.material}。{ctx.detail_material}"),
        ("03 边缘细节", ctx.detail_fringe),
        ("04 尺寸价格", "结合登记表直接给出选择参考。"),
        ("05 使用建议", ctx.detail_scene),
    ]

    for idx, (title, body) in enumerate(sections):
        top = idx * 1200
        draw.rounded_rectangle((28, top + 24, page_w - 28, top + 1176), radius=30, fill="#FFFDF9")
        hero = crop_cover(edge if idx in {0, 2} else texture, (650, 500), focus_bottom=idx in {0, 2})
        canvas.paste(hero, (50, top + 90))
        draw.text((54, top + 626), title, font=load_font(42), fill="#1F2937")
        y = draw_multiline(draw, body, (54, top + 686), load_font(26), "#6B7280", max_chars=18)

        if idx == 3:
            y += 20
            for size, price in zip(ctx.sizes, ctx.prices):
                draw.rounded_rectangle((54, y, 696, y + 70), radius=18, fill="#F3EEE7")
                draw.text((82, y + 18), size, font=load_font(24), fill="#374151")
                draw.text((562, y + 18), price, font=load_font(24), fill="#111827")
                y += 86
        elif idx == 4:
            tips = [
                "铺拍时把流苏边露出来，不要只拍布中心。",
                "主图先讲风格，不先堆防水防油参数。",
                "近拍图一定保留布纹，别磨皮磨到没材质。",
            ]
            for tip in tips:
                draw.text((70, y), f"• {tip}", font=load_font(24), fill="#4B5563")
                y += 42

    master = output_dir / "05_详情页_01.jpg"
    canvas.save(master, quality=95)

    slices: list[Path] = []
    for idx in range(5):
        top = idx * 1200
        part = canvas.crop((0, top, page_w, top + 1200))
        path = output_dir / f"05_详情页_切片_{idx + 1:02d}.jpg"
        part.save(path, quality=95)
        slices.append(path)
    return master, slices


def update_workbook(ctx: ProductContext) -> None:
    wb = load_workbook(ctx.workbook_path)
    ws = wb["SKU信息登记"]
    ws["B2"] = ctx.title
    ws["B3"] = ctx.guide_title
    ws["B6"] = "、".join(ctx.style_tags)
    wb.save(ctx.workbook_path)


def run(product_name: str | None = None) -> dict[str, list[Path] | Path]:
    ctx = read_context(product_name)
    texture = Image.open(ctx.texture_image).convert("RGB")
    edge = clean_edge_source(Image.open(ctx.edge_image).convert("RGB"))

    white_dir = ensure_output_dir(ctx.product_dir, "白底图", "01_白底图")
    square_dir = ensure_output_dir(ctx.product_dir, "1比1主图", "02_1比1主图")
    portrait_dir = ensure_output_dir(ctx.product_dir, "3比4主图", "03_3比4主图")
    detail_dir = ensure_output_dir(ctx.product_dir, "详情页", "05_详情页")

    white_path = white_dir / "01_白底图_01.jpg"
    make_white_bg(texture, edge, white_path)
    square_paths = render_square_series(ctx, texture, edge, square_dir)
    portrait_paths = render_portrait_series(ctx, texture, edge, portrait_dir)
    detail_master, detail_slices = render_detail_pages(ctx, texture, edge, detail_dir)

    update_workbook(ctx)

    return {
        "white": [white_path],
        "square": square_paths,
        "portrait": portrait_paths,
        "detail": [detail_master, *detail_slices],
        "workbook": [ctx.workbook_path],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="桌布自动生图工作流试跑脚本")
    parser.add_argument("--product", help="产品子文件夹名，例如 商品1")
    args = parser.parse_args()

    outputs = run(args.product)
    for key, paths in outputs.items():
        print(f"[{key}]")
        for path in paths:
            print(path)


if __name__ == "__main__":
    main()
