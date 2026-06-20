"""运行此脚本生成商品信息模板 Excel"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path


def create_template(output_path: str = "data/templates/商品信息模板.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "商品信息"

    # 必填列（红色星号标记）
    required = {'序号', '商品标题', '颜色', '尺寸', '价格'}

    headers = [
        # ── 基础信息 ──
        '序号', '商品标题', '颜色', '尺寸', '价格', '库存',
        # ── 物流 ──
        '运费模板', '发货天数',
        # ── 商品属性（选填，有值才填入淘宝对应字段）──
        '类目ID',
        '表面材质', '适用场景', '形状', '主图案类型',
        '产地', '风格', '工艺类型', '功能',
        '清洗方式', '适用季节', '适用桌型',
        # ── 多件优惠 ──
        '多件优惠', '折扣率',
        # ── 其他 ──
        '详情描述',
    ]

    col_widths = [
        8, 45, 12, 15, 10, 10,
        22, 10,
        14,
        12, 12, 10, 14,
        12, 10, 10, 10,
        10, 10, 12,
        10, 8,
        40,
    ]

    # 样式
    hdr_font      = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill      = PatternFill(fill_type='solid', fgColor='2E75B6')
    attr_fill     = PatternFill(fill_type='solid', fgColor='1F6B3A')  # 属性列深绿
    extra_fill    = PatternFill(fill_type='solid', fgColor='7B3F9E')  # 优惠列紫色
    align_ctr     = Alignment(horizontal='center', vertical='center')
    align_lft     = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin          = Side(style='thin', color='CCCCCC')
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 属性列范围
    attr_start = headers.index('类目ID') + 1
    attr_end   = headers.index('适用桌型') + 1
    extra_start = headers.index('多件优惠') + 1
    extra_end   = headers.index('折扣率') + 1

    # 写表头
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = hdr_font
        cell.alignment = align_ctr
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = w

        if attr_start <= col_idx <= attr_end:
            cell.fill = attr_fill
        elif extra_start <= col_idx <= extra_end:
            cell.fill = extra_fill
        else:
            cell.fill = hdr_fill

    ws.row_dimensions[1].height = 22

    # 示例数据
    rows = [
        ['001', '田园风碎花印花桌布纯棉环保儿童房书桌茶几装饰生日派对拍照神器',
         '彩色格纹', '60cm*60cm',  '39', '100',
         '光阴童话', '2',
         '121458013',
         '棉质', '日常使用', '矩形', '碎花图案',
         '中国大陆', '法式', '印花', '透气耐洗',
         '', '', '',
         '是', '9.5',
         ''],
        ['001', '', '彩色格纹', '100cm*140cm', '50', '100',
         '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
        ['001', '', '彩色格纹', '180cm*140cm', '63', '100',
         '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
    ]

    alt_fill = PatternFill(fill_type='solid', fgColor='F2F7FC')
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            h = headers[col_idx - 1]
            cell.alignment = align_lft if h in ('商品标题', '详情描述') else align_ctr
            cell.border    = border
        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = 'A2'

    # 填写说明 sheet
    ws2 = wb.create_sheet('填写说明')
    notes = [
        ('字段',      '必填', '说明'),
        ('序号',      '是',   '商品编号，与图片文件夹名一致（如 001）。同商品所有 SKU 行写相同序号。'),
        ('商品标题',  '是',   '最多 30个汉字（60字符）。首行填写，同序号后续行留空。'),
        ('颜色',      '是',   '每行填一种颜色，如"彩色格纹"。'),
        ('尺寸',      '是',   '每行填一种尺寸，如"60cm*60cm"。'),
        ('价格',      '是',   '该颜色+尺寸组合的售价，纯数字。'),
        ('库存',      '否',   '该 SKU 库存数量，默认 100。'),
        ('运费模板',  '是',   '首行填写，需与淘宝店铺运费模板名称完全一致。'),
        ('发货天数',  '是',   '0=今日发, 1=24h内, 2=48h内(默认), 3=大于48h。'),
        ('类目ID',    '否',   '淘宝类目ID，如桌布=121458013。填写后可直链跳过类目弹窗。'),
        ('表面材质',  '否',   '如：棉质、涤纶、棉麻（下拉选择，需与淘宝选项完全一致）。'),
        ('适用场景',  '否',   '如：日常使用、户外野餐（下拉选择）。'),
        ('形状',      '否',   '如：矩形、圆形、椭圆形（可自由输入）。'),
        ('主图案类型','否',   '如：碎花图案、纯色、格纹（下拉选择）。'),
        ('产地',      '否',   '如：中国大陆（下拉选择，默认不填为中国大陆）。'),
        ('风格',      '否',   '如：法式、北欧、现代简约（下拉选择）。'),
        ('工艺类型',  '否',   '如：印花、刺绣、编织（可自由输入）。'),
        ('功能',      '否',   '如：透气耐洗、防水（可自由输入）。'),
        ('清洗方式',  '否',   '如：机洗（下拉选择）。'),
        ('适用季节',  '否',   '如：四季（下拉选择）。'),
        ('适用桌型',  '否',   '如：方桌（下拉选择）。'),
        ('多件优惠',  '是',   '"是"或"否"，淘宝新版必填，建议填"是"。'),
        ('折扣率',    '否',   '多件优惠折扣，如"9.5"表示9.5折，范围5.0~9.9。'),
        ('详情描述',  '否',   '商品文字描述，图片通过详情图文件夹上传。'),
    ]
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 6
    ws2.column_dimensions['C'].width = 72

    hdr2_fill = PatternFill(fill_type='solid', fgColor='2E75B6')
    for r, row in enumerate(notes, start=1):
        for c, v in enumerate(row, start=1):
            cell = ws2.cell(row=r, column=c, value=v)
            if r == 1:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = hdr2_fill
            elif c == 1:
                cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical='center', wrap_text=(c == 3))
        ws2.row_dimensions[r].height = 20 if r > 1 else 22

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"模板已生成: {output_path}")


if __name__ == '__main__':
    create_template()
