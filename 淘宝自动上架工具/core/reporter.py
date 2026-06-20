"""
任务结束后生成 Excel 报告。
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from pathlib import Path
from loguru import logger
from models.product import Product, TaskStatus


def generate_report(products: list[Product], output_dir: str | Path = 'data') -> Path:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
    path = output_dir / f'上架报告_{ts}.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '上架结果'

    # ── 表头 ──────────────────────────────────────────────────────────────
    headers    = ['序号', '商品标题', 'SKU数', '状态', '商品ID', '失败原因']
    col_widths = [8, 45, 8, 10, 18, 40]

    thin   = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill(fill_type='solid', fgColor='2E75B6')
    ctr      = Alignment(horizontal='center', vertical='center')
    lft      = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = ctr; c.border = border
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 22

    # ── 数据行 ────────────────────────────────────────────────────────────
    fill_ok   = PatternFill(fill_type='solid', fgColor='E2EFDA')
    fill_fail = PatternFill(fill_type='solid', fgColor='FCE4D6')
    fill_skip = PatternFill(fill_type='solid', fgColor='F2F2F2')

    status_label = {
        TaskStatus.DONE:    '✅ 成功',
        TaskStatus.FAILED:  '❌ 失败',
        TaskStatus.PENDING: '⏳ 未执行',
        TaskStatus.SKIPPED: '⏭ 跳过',
        TaskStatus.RUNNING: '⏸ 中断',
    }

    for ri, p in enumerate(products, 2):
        fill = (fill_ok   if p.status == TaskStatus.DONE
                else fill_fail if p.status == TaskStatus.FAILED
                else fill_skip)
        row_data = [p.seq, p.title, len(p.skus),
                    status_label.get(p.status, p.status.name),
                    p.item_id, p.error]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = fill; c.border = border
            c.alignment = lft if ci in (2, 6) else ctr
        ws.row_dimensions[ri].height = 18

    # ── 汇总行 ────────────────────────────────────────────────────────────
    total   = len(products)
    done    = sum(1 for p in products if p.status == TaskStatus.DONE)
    failed  = sum(1 for p in products if p.status == TaskStatus.FAILED)
    pending = total - done - failed

    sum_row = total + 2
    ws.cell(row=sum_row, column=1, value='汇总').font = Font(bold=True)
    ws.cell(row=sum_row, column=2,
            value=f'共 {total} 个  ✅成功 {done}  ❌失败 {failed}  ⏳未执行 {pending}')

    ws.freeze_panes = 'A2'
    wb.save(str(path))
    logger.info(f"报告已生成: {path}  (成功{done}/失败{failed}/共{total})")
    return path
