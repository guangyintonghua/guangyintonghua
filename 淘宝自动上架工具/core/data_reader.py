# -*- coding: utf-8 -*-
"""
商品数据读取：支持两种格式
  1. 标准横向表格（create_template 生成的模板）
  2. 竖向键值对（SKU信息记录表.xlsx，每行一个字段）

导入文件夹时，工具自动识别：
  - 文件夹本身含 Excel → 当作单个产品
  - 文件夹含多个子文件夹各含 Excel → 扫描所有子文件夹
"""
from pathlib import Path
from loguru import logger
import openpyxl

from models.product import Product, SkuRow


_REQUIRED_COLS = {'序号', '商品标题', '颜色', '尺寸', '价格'}

_ATTRIBUTE_COLS = {
    '表面材质', '适用场景', '形状', '主图案类型', '产地',
    '风格', '功能', '工艺类型', '清洗方式', '适用季节', '适用桌型',
    '底层材质', '防滑性能', '固定方式', '定制服务',
}


# ── 文件夹导入入口（对外主接口）────────────────────────────────────────────────

def scan_and_read_folder(folder: str | Path) -> list[Product]:
    """
    扫描文件夹，自动识别并读取所有产品。

    逻辑：
      1. 先找本文件夹内的 Excel → 若有，读它（横向表格 or 单竖向产品）
      2. 再扫描所有子文件夹，每个含 Excel 的子文件夹 = 一个产品
      所有结果合并返回，按文件夹名排序。
    """
    folder = Path(folder)
    if not folder.is_dir():
        raise NotADirectoryError(f"不是有效文件夹: {folder}")

    products: list[Product] = []

    # 第一步：本文件夹内的 Excel
    root_excel = _find_excel(folder)
    if root_excel:
        from core.sku_reader import is_sku_sheet_format
        if is_sku_sheet_format(root_excel):
            from core.sku_reader import read_sku_sheet
            p = read_sku_sheet(folder, seq=folder.name)
            if p:
                products.append(p)
        else:
            # 标准横向表格 → 可能含多个产品
            try:
                ps = read_excel(root_excel)
                attach_images(ps, folder)
                products.extend(ps)
            except Exception as e:
                logger.warning(f"读取横向表格失败: {e}")

    # 第二步：子文件夹（每个含 Excel 的子文件夹 = 一个产品）
    seq_num = len(products) + 1
    for subdir in sorted(d for d in folder.iterdir() if d.is_dir()):
        sub_excel = _find_excel(subdir)
        if not sub_excel:
            continue
        from core.sku_reader import is_sku_sheet_format, read_sku_sheet
        if is_sku_sheet_format(sub_excel):
            p = read_sku_sheet(subdir, seq=subdir.name)
        else:
            try:
                ps = read_excel(sub_excel)
                attach_images(ps, subdir)
                products.extend(ps)
                seq_num += len(ps)
                continue
            except Exception as e:
                logger.warning(f"[{subdir.name}] 读取失败: {e}")
                continue
        if p:
            products.append(p)
            seq_num += 1

    if not products:
        raise FileNotFoundError(
            f"文件夹中未找到可识别的商品数据\n"
            f"路径: {folder}\n\n"
            f"支持的格式：\n"
            f"  · 含「SKU信息记录表.xlsx」的产品文件夹\n"
            f"  · 含标准商品信息模板（序号/商品标题/颜色/尺寸/价格 列）的 Excel"
        )

    logger.info(f"导入完成: 共 {len(products)} 个商品")
    return products


# ── 兼容旧接口 ────────────────────────────────────────────────────────────────

def scan_folder(folder: str | Path) -> Path:
    """找到文件夹内第一个 Excel 文件路径（供旧代码调用）"""
    xl = _find_excel(Path(folder))
    if not xl:
        raise FileNotFoundError(f"文件夹中未找到 Excel 文件: {folder}")
    return xl


def _find_excel(folder: Path) -> Path | None:
    """返回文件夹内第一个非临时 Excel，无则返回 None"""
    for ext in ('*.xlsx', '*.xls'):
        for f in sorted(folder.glob(ext)):
            if not f.name.startswith('~$'):
                return f
    return None


# ── 标准横向表格读取 ──────────────────────────────────────────────────────────

def read_excel(excel_path: str | Path) -> list[Product]:
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"找不到 Excel 文件: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    headers = [str(c.value).strip() if c.value else ''
               for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx for idx, name in enumerate(headers)}

    missing = _REQUIRED_COLS - set(col.keys())
    if missing:
        raise ValueError(
            f"Excel 缺少必需列: {missing}\n"
            f"当前列: {[h for h in headers if h]}\n\n"
            f"请使用「python create_template.py」生成标准模板，\n"
            f"或将文件夹内的「SKU信息记录表.xlsx」改用产品文件夹模式导入。"
        )

    def get(row, name: str, default=''):
        idx = col.get(name)
        if idx is None:
            return default
        v = row[idx].value
        return str(v).strip() if v is not None else default

    products: dict[str, Product] = {}
    last_seq = None

    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        seq = get(row, '序号')
        if seq:
            last_seq = seq
        elif last_seq:
            seq = last_seq
        else:
            continue

        if seq not in products:
            title = get(row, '商品标题')
            if not title:
                logger.warning(f"第 {row_num} 行序号 {seq!r} 缺少商品标题，跳过")
                continue

            char_count = sum(2 if '一' <= c <= '鿿' else 1 for c in title)
            if char_count > 60:
                logger.warning(f"第 {row_num} 行标题超长 {char_count}/60")

            discount_raw     = get(row, '多件优惠', '是').strip().lower()
            discount_enabled = discount_raw not in ('否', 'no', 'false', '0')

            p = Product(
                seq=seq,
                title=title,
                cat_id=get(row, '类目ID', ''),
                shipping_tpl=get(row, '运费模板', '光阴童话'),
                delivery_days=int(get(row, '发货天数', '2') or 2),
                description=get(row, '详情描述', ''),
                multi_discount_enabled=discount_enabled,
                multi_discount_rate=get(row, '折扣率', '9.5') or '9.5',
            )
            for attr_name in _ATTRIBUTE_COLS:
                val = get(row, attr_name, '')
                if val:
                    p.attributes[attr_name] = val
            products[seq] = p

        p = products[seq]
        color     = get(row, '颜色')
        size      = get(row, '尺寸')
        price_str = get(row, '价格', '0')
        stock_str = get(row, '库存', '100')

        try:
            price = float(price_str)
            stock = int(float(stock_str))
        except ValueError:
            logger.warning(f"第 {row_num} 行价格/库存格式错误，跳过")
            continue

        if not color or not size:
            logger.warning(f"第 {row_num} 行颜色/尺寸为空，跳过")
            continue

        p.skus.append(SkuRow(color=color, size=size, price=price, stock=stock))

    result = [p for p in products.values() if p.skus]
    logger.info(f"读取完成: {len(result)} 个商品，"
                f"{sum(len(p.skus) for p in result)} 个SKU")
    return result


# ── 图片关联（横向表格用） ────────────────────────────────────────────────────

def attach_images(products: list[Product], root_dir: str | Path) -> None:
    """将产品与图片文件夹关联（横向表格模式用）"""
    root = Path(root_dir)
    subdirs = {d.name: d for d in root.iterdir() if d.is_dir()}
    for p in products:
        d = _find_image_dir(p.seq, subdirs)
        if d:
            p.image_dir = d
        else:
            logger.warning(f"序号 {p.seq} 的图片文件夹未找到")


def _find_image_dir(seq: str, subdirs: dict[str, Path]) -> Path | None:
    if seq in subdirs:
        return subdirs[seq]
    for name, d in subdirs.items():
        if name.startswith(seq) and (
            len(name) == len(seq) or name[len(seq)] in ('_', ' ', '-', '.')
        ):
            return d
    for name, d in subdirs.items():
        if seq in name:
            return d
    return None
