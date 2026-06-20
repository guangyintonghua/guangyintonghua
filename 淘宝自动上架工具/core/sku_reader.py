# -*- coding: utf-8 -*-
"""
读取竖向键值对格式的 SKU 信息登记表（SKU信息记录表/登记表.xlsx）。

格式：第一行为空，从第二行起 A列=字段名，B/C/D列=值（多值横向排列）。

  行02: 商品标题 | ins风田园风碎花...
  行07: 颜色     | 彩色格纹
  行09: 产品尺寸 | 60cm*60cm | 100cm*140cm | 180cm*140cm
  行10: 价格     | 39元      | 50元        | 63元
  行11: 库存     | 100       | 100         | 100
  行29: 主图目录 | E:\\桌布\\5款\\趣味印花桌布\\主图
  行30: 详情图目录| E:\\桌布\\5款\\趣味印花桌布\\详情图
"""
from pathlib import Path
from loguru import logger
import openpyxl

from models.product import Product, SkuRow


# 字段名 → Product.attributes 键名
# 键名尽量与淘宝表单上的标签一致，方便 _fill_attr_by_label 查找
_ATTR_MAP = {
    '表面材质':       '表面材质',   # Excel列名→Taobao属性键
    '面料材质':       '表面材质',
    '材质':           '表面材质',
    '工艺类型':       '工艺类型',
    '工艺描述':       '工艺类型',
    '清洗方式':       '清洗方式',
    '适用场景':       '适用场景',
    '形状':           '形状',
    '主图案类型':     '主图案类型',
    '图案类型':       '主图案类型',
    '产地':           '产地',
    '风格标签':       '风格',
    '风格':           '风格',
    '使用桌型':       '使用桌型',
    '适用桌型':       '使用桌型',
    # '厚度' 故意不映射：Taobao表单中对应'台板厚度'（板材厚度），桌布不适用
    '品牌':           '品牌',
    '功能':           '功能',
    '使用功能':       '功能',
    '颜色系列':       '颜色系列',
}

# 商品标题字段的所有可能名称
_TITLE_KEYS  = ('商品标题', '产品名称', '货品名称', '标题', '品名', '品名称')
# 导购标题（短标题）字段的所有可能名称
_GUIDE_KEYS  = ('导购标题', '短标题', '推广标题', '导购')
# 尺寸字段的所有可能名称
_SIZE_KEYS   = ('产品尺寸', '商品尺寸', '规格', '尺寸', '型号')
# 颜色字段
_COLOR_KEYS  = ('颜色', '色系', '颜色分类')


def is_sku_sheet_format(xl_path: Path) -> bool:
    """判断 Excel 是否为竖向键值对格式（第一行空、第二行首列有字段名）"""
    try:
        wb   = openpyxl.load_workbook(xl_path, data_only=True, read_only=True)
        ws   = wb.active
        rows = list(ws.iter_rows(max_row=3, values_only=True))
        wb.close()
        if len(rows) < 2:
            return False
        first_empty = all(v is None or str(v).strip() == '' for v in rows[0])
        second_has  = rows[1][0] is not None and str(rows[1][0]).strip() != ''
        return first_empty and second_has
    except Exception:
        return False


def read_sku_sheet(folder: Path, seq: str = '') -> Product | None:
    """
    读取产品文件夹内的竖向 Excel，返回 Product 对象。
    folder: 产品文件夹路径（含 Excel、主图/、详情图/）
    seq:    序号标识（默认用文件夹名）
    """
    xl_files = list(folder.glob('*.xlsx')) + list(folder.glob('*.xls'))
    if not xl_files:
        return None

    # 排除临时文件
    xl_files = [f for f in xl_files if not f.name.startswith('~$')]
    if not xl_files:
        return None

    xl_path = xl_files[0]
    seq     = seq or folder.name

    try:
        wb   = openpyxl.load_workbook(xl_path, data_only=True)
        ws   = wb.active
        data = _parse_kv(ws)
    except Exception as e:
        logger.error(f"[{seq}] 读取 {xl_path.name} 失败: {e}")
        return None

    # ── 标题 ──────────────────────────────────────────────────────────────
    title = ''
    for k in _TITLE_KEYS:
        title = data.get(k)
        if title:
            break
    if not title:
        logger.warning(f"[{seq}] 未找到商品标题字段（尝试了: {_TITLE_KEYS}）"
                       f"\n  实际字段: {data.keys()}")
        return None

    # ── 导购标题 ──────────────────────────────────────────────────────────
    guide_title = ''
    for k in _GUIDE_KEYS:
        guide_title = data.get(k)
        if guide_title:
            break

    # ── 颜色 ──────────────────────────────────────────────────────────────
    color = ''
    for k in _COLOR_KEYS:
        color = data.get(k)
        if color:
            break
    if not color:
        color = '默认'

    # ── 尺寸 / 价格 / 库存 ───────────────────────────────────────────────
    sizes: list[str] = []
    for k in _SIZE_KEYS:
        sizes = data.get_list(k)
        if sizes:
            break

    prices_raw = data.get_list('价格') or data.get_list('售价')
    stocks_raw = data.get_list('库存') or data.get_list('数量')

    prices = _parse_numbers(prices_raw)
    stocks = _parse_ints(stocks_raw, default=100)

    if not sizes:
        logger.warning(f"[{seq}] 未找到尺寸字段（尝试了: {_SIZE_KEYS}）")
        return None
    if not prices:
        logger.warning(f"[{seq}] 未找到价格字段")
        return None

    # ── SKU 列表 ──────────────────────────────────────────────────────────
    skus: list[SkuRow] = []
    for i, size in enumerate(sizes):
        price = prices[i] if i < len(prices) else prices[-1]
        stock = stocks[i] if i < len(stocks) else 100
        skus.append(SkuRow(color=color, size=size, price=price, stock=stock))

    # ── 属性 ──────────────────────────────────────────────────────────────
    attributes: dict[str, str] = {}
    for src_key, dst_key in _ATTR_MAP.items():
        val = data.get(src_key)
        if val and dst_key not in attributes:
            attributes[dst_key] = val

    # ── 图片目录（优先用文件夹内的 主图/ 详情图/，否则用 Excel 内记录的路径）──
    img_dir = folder   # image_processor 会在此目录内找 主图/ 详情图/
    # 若本地没有，尝试 Excel 记录的绝对路径
    if not (folder / '主图').is_dir():
        abs_main = data.get('主图目录')
        if abs_main and Path(abs_main).is_dir():
            img_dir = Path(abs_main).parent   # 上级即产品目录

    # ── 组装 Product ──────────────────────────────────────────────────────
    p = Product(
        seq=seq,
        title=title,
        guide_title=guide_title,
        skus=skus,
        attributes=attributes,
        image_dir=img_dir,
        shipping_tpl='光阴童话',
    )

    # 若运费字段含"包邮"就默认光阴童话
    shipping_raw = data.get('运费', '')
    if shipping_raw and '包邮' not in shipping_raw:
        p.shipping_tpl = shipping_raw

    logger.info(f"[{seq}] 读取完成: 《{title[:20]}》  "
                f"{len(skus)} 个SKU  图片目录: {img_dir}")
    return p


# ── 内部辅助 ──────────────────────────────────────────────────────────────────

class _KVStore:
    def __init__(self):
        self._d: dict[str, list[str]] = {}

    def add(self, key: str, values: list[str]):
        if key not in self._d:        # 只保留第一次出现的值
            self._d[key] = values

    def get(self, key: str, default: str = '') -> str:
        vals = self._d.get(key, [])
        return vals[0] if vals else default

    def get_list(self, key: str) -> list[str]:
        return self._d.get(key, [])

    def keys(self) -> list[str]:
        return list(self._d.keys())


def _parse_kv(ws) -> _KVStore:
    store = _KVStore()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        key    = str(row[0]).strip()
        values = [str(v).strip() for v in row[1:] if v is not None and str(v).strip()]
        if key and values:
            store.add(key, values)
    return store


def _parse_numbers(raw: list[str]) -> list[float]:
    result = []
    for v in raw:
        cleaned = v.replace('元', '').replace('￥', '').replace(',', '').strip()
        try:
            result.append(float(cleaned))
        except ValueError:
            pass
    return result


def _parse_ints(raw: list[str], default: int = 100) -> list[int]:
    result = []
    for v in raw:
        try:
            result.append(int(float(v)))
        except (ValueError, TypeError):
            result.append(default)
    return result
